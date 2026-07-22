"""Per-molecule state analysis.

Computes, per molecule and per conformer-selection mode:

  - Triplet energy             ΔE(T1) = E_comb(T1) − E_comb(S0)
  - T1 reorganisation energy   full 4-point Marcus on the S0↔T1 surface,
                               requires e_vert_spin_change on BOTH sides
  - Redox free energies        ΔG_ox = G(ox) − G(S0), ΔG_red = G(S0) − G(red)
  - E vs SCE in MeCN           ΔG/F − E_ref_SCE  (only when MeCN solvation
                               is detected in any header)
  - Redox reorganisation       4-point Marcus on each S0↔ox / S0↔red surface,
                               requires the appropriate vertical SP on
                               both ends

Two conformer-selection modes:

  - ``lowest_energy``: per state, pick the conformer with the smallest
    free-energy-corrected E_combined.
  - ``rmsd_matched``: S0 keeps its lowest-energy conformer; every other
    state picks the conformer with the smallest Kabsch RMSD relative to
    the S0 reference geometry. Direct atom-by-atom RMSD is appropriate
    because every state starts from the same conformer seeds, so the
    optimised geometries preserve atom ordering.

Reference values:

  - ``SCE_ABS_MECN = 4.42 V`` (Pavlishchuk & Addison, *Inorg. Chim. Acta*
    2000, 298, 97-102).
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import numpy as np
from sqlmodel import col, func, select

from autodft.db import get_session
from autodft.extraction.extractor import ConformerResult, PipelineExtractor
from autodft.models import (
    ComputationHeader,
    ComputationTask,
    Molecule,
    MoleculeState,
)
from autodft.models.geometry import MoleculeGeometry


HARTREE_TO_EV = 27.211386245988
HARTREE_TO_KCAL_PER_MOL = 627.5094740631
SCE_ABS_MECN = 4.42  # V, absolute SCE potential in MeCN (Pavlishchuk & Addison 2000)


# Match MeCN solvation written in any of the common ORCA forms:
#   "! CPCM(Acetonitrile)" / "! SMD(MeCN)" / "%cpcm  smd true  SMDsolvent \"acetonitrile\" end"
#   "! SMD(CH3CN)" is accepted by ORCA too and was previously missed.
_MECN_NAMES = r"(?:acetonitrile|mecn|ch3cn|acetonitril)"
_MECN_RE = re.compile(
    rf"(CPCM|SMD)\s*\(\s*{_MECN_NAMES}\s*\)"
    rf"|smdsolvent\s*[=\s]*[\"']?\s*{_MECN_NAMES}"
    rf"|solvent\s*[=\s]+[\"']?\s*{_MECN_NAMES}",
    re.IGNORECASE,
)


def detects_mecn_solvation(header_text: Optional[str]) -> bool:
    return bool(header_text and _MECN_RE.search(header_text))


# ----------------------------------------------------------------------
# Geometry parsing + Kabsch RMSD
# ----------------------------------------------------------------------


def parse_xyz(xyz_data: str) -> tuple[list[str], np.ndarray]:
    """Parse an XYZ string into element symbols + (N, 3) coords.

    Accepts both the standard ``.xyz`` layout (count + comment + rows)
    and a bare ``symbol x y z`` table.
    """
    if not xyz_data:
        return [], np.empty((0, 3))
    lines = xyz_data.strip().splitlines()
    head = lines[0].strip().split() if lines else []
    if len(head) == 1 and head[0].isdigit():
        # Standard .xyz: count line + comment line + atom rows.
        lines = lines[2:] if len(lines) >= 2 else []
    elems: list[str] = []
    coords: list[list[float]] = []
    for ln in lines:
        parts = ln.split()
        if len(parts) < 4:
            continue
        try:
            coords.append([float(parts[1]), float(parts[2]), float(parts[3])])
        except ValueError:
            continue
        elems.append(parts[0])
    if not coords:
        return elems, np.empty((0, 3))
    return elems, np.asarray(coords, dtype=float)


def kabsch_rmsd(p: np.ndarray, q: np.ndarray) -> float:
    """Kabsch-aligned RMSD in Å between two (N, 3) coordinate arrays.

    Caller must ensure both arrays share atom ordering — no atom mapping
    is performed.
    """
    if p.shape != q.shape or p.size == 0:
        raise ValueError("coordinate arrays must share a non-empty shape")
    p_c = p - p.mean(axis=0)
    q_c = q - q.mean(axis=0)
    h = p_c.T @ q_c
    u, _, vt = np.linalg.svd(h)
    d = float(np.sign(np.linalg.det(vt.T @ u.T)))
    diag = np.diag([1.0, 1.0, d])
    rot = vt.T @ diag @ u.T
    rotated = p_c @ rot.T
    diff = rotated - q_c
    return float(np.sqrt((diff * diff).sum() / p.shape[0]))


# ----------------------------------------------------------------------
# Per-mode analysis
# ----------------------------------------------------------------------


@dataclass
class StatePick:
    state: str
    conformer_index: int
    opt_task_id: int
    rmsd_to_s0: Optional[float]  # Å, None for S0 and when RMSD can't be computed
    e_sp: Optional[float]                # Hartree
    e_correction: Optional[float]        # Hartree
    e_combined: Optional[float]          # Hartree
    e_vert_ox: Optional[float]           # Hartree
    e_vert_red: Optional[float]          # Hartree
    e_vert_spin_change: Optional[float]  # Hartree


@dataclass
class ModeAnalysis:
    """Energies are reported in Hartree; the frontend converts to the
    user's chosen unit. Redox potentials stay in V (since the user fixed
    that on the UI side)."""
    mode: str
    picks: dict[str, StatePick]
    # Triplet: adiabatic gap + forward / backward inner reorganisation.
    triplet_dE_adiabatic_h: Optional[float] = None
    triplet_lambda_fwd_h: Optional[float] = None
    triplet_lambda_bwd_h: Optional[float] = None
    # Oxidation: adiabatic free energy, E vs SCE, forward / backward λ.
    ox_dG_h: Optional[float] = None
    ox_E_vs_sce_v: Optional[float] = None
    ox_lambda_fwd_h: Optional[float] = None
    ox_lambda_bwd_h: Optional[float] = None
    # Reduction: same layout as ox.
    red_dG_h: Optional[float] = None
    red_E_vs_sce_v: Optional[float] = None
    red_lambda_fwd_h: Optional[float] = None
    red_lambda_bwd_h: Optional[float] = None


def _compute_derived(picks: dict[str, StatePick], mecn: bool) -> ModeAnalysis:
    """Compute every derived quantity in Hartree (energies) and V
    (redox potentials).

    Reorganisation convention — for an electron transfer R → P:
        λ_fwd = E_P(R_geom) − E_P(P_geom)       (cost on the PRODUCT surface)
        λ_bwd = E_R(P_geom) − E_R(R_geom)       (cost on the REACTANT surface)
    The total Marcus reorg is (λ_fwd + λ_bwd) / 2 in the harmonic limit;
    here we surface both halves separately so asymmetries are visible.
    """
    ana = ModeAnalysis(mode="", picks=picks)
    s0 = picks.get("S0")
    t1 = picks.get("T1")
    ox = picks.get("ox")
    red = picks.get("red")

    # ── Triplet ───────────────────────────────────────────────────────
    if s0 and t1 and s0.e_combined is not None and t1.e_combined is not None:
        ana.triplet_dE_adiabatic_h = t1.e_combined - s0.e_combined

    # λ_T1_fwd: on T1 surface = E_vert_spin_change(S0) − E_sp(T1)
    if s0 and t1 and s0.e_vert_spin_change is not None and t1.e_sp is not None:
        ana.triplet_lambda_fwd_h = s0.e_vert_spin_change - t1.e_sp
    # λ_T1_bwd: on S0 surface = E_vert_spin_change(T1) − E_sp(S0)
    if s0 and t1 and t1.e_vert_spin_change is not None and s0.e_sp is not None:
        ana.triplet_lambda_bwd_h = t1.e_vert_spin_change - s0.e_sp

    # ── Oxidation ─────────────────────────────────────────────────────
    if s0 and ox and s0.e_combined is not None and ox.e_combined is not None:
        ana.ox_dG_h = ox.e_combined - s0.e_combined
        if mecn:
            ana.ox_E_vs_sce_v = ana.ox_dG_h * HARTREE_TO_EV - SCE_ABS_MECN

    # λ_ox_fwd: on ox surface = E_vert_ox(S0) − E_sp(ox)
    if s0 and ox and s0.e_vert_ox is not None and ox.e_sp is not None:
        ana.ox_lambda_fwd_h = s0.e_vert_ox - ox.e_sp
    # λ_ox_bwd: on S0 surface = E_vert_red(ox) − E_sp(S0)
    if s0 and ox and ox.e_vert_red is not None and s0.e_sp is not None:
        ana.ox_lambda_bwd_h = ox.e_vert_red - s0.e_sp

    # ── Reduction ─────────────────────────────────────────────────────
    # ΔG_red = G(S0) − G(red): positive when reduction is unfavourable,
    # matches the sign convention used by the SCE formula below
    # (E°_red,SCE = ΔG/F − E_ref).
    if s0 and red and s0.e_combined is not None and red.e_combined is not None:
        ana.red_dG_h = s0.e_combined - red.e_combined
        if mecn:
            ana.red_E_vs_sce_v = ana.red_dG_h * HARTREE_TO_EV - SCE_ABS_MECN

    # λ_red_fwd: on red surface = E_vert_red(S0) − E_sp(red)
    if s0 and red and s0.e_vert_red is not None and red.e_sp is not None:
        ana.red_lambda_fwd_h = s0.e_vert_red - red.e_sp
    # λ_red_bwd: on S0 surface = E_vert_ox(red) − E_sp(S0)
    if s0 and red and red.e_vert_ox is not None and s0.e_sp is not None:
        ana.red_lambda_bwd_h = red.e_vert_ox - s0.e_sp

    return ana


# ----------------------------------------------------------------------
# Project-level driver
# ----------------------------------------------------------------------


def _parse_float(s: Optional[str]) -> Optional[float]:
    if s is None or s == "" or s == "None":
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _load_results_from_archive_csv(csv_path: Path) -> list[ConformerResult]:
    """Reconstruct ConformerResult rows from a frozen archive CSV.

    The archive step writes <export>/<project>/<project>.csv with every
    column we need for state analysis (all energies in Hartree, plus
    opt_task_id so we can still resolve geometries from the DB for
    Kabsch RMSD). Used when output.out files have been wiped from
    comp_data/.
    """
    results: list[ConformerResult] = []
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                mol_id = int(row["molecule_id"])
                conf_idx = int(row["conformer_index"])
                opt_task_id = int(row["opt_task_id"])
            except (KeyError, TypeError, ValueError):
                continue
            results.append(ConformerResult(
                molecule_id=mol_id,
                smiles=row.get("smiles", ""),
                state=row.get("state", ""),
                conformer_index=conf_idx,
                opt_task_id=opt_task_id,
                e_singlepoint=_parse_float(row.get("e_singlepoint")),
                e_correction=_parse_float(row.get("e_correction")),
                e_combined=_parse_float(row.get("e_combined")),
                e_vert_spin_change=_parse_float(row.get("e_vert_spin_change")),
                e_vert_ox=_parse_float(row.get("e_vert_ox")),
                e_vert_red=_parse_float(row.get("e_vert_red")),
            ))
    return results


# Cached analyses, keyed by project name. Each entry is
# (signature, payload); the signature is a cheap pair of aggregates that
# changes whenever any task in the project is updated, so a stale cache
# cannot outlive new results.
_ANALYSIS_CACHE: dict[str, tuple[tuple, dict]] = {}


def _cache_signature(project_name: str) -> tuple:
    """A cheap fingerprint of the project's current results."""
    with get_session() as session:
        molecules = session.exec(
            select(func.count()).select_from(Molecule)
            .where(Molecule.project_name == project_name)
        ).one()
        tasks, latest = session.exec(
            select(func.count(), func.max(ComputationTask.updated_at))
            .select_from(ComputationTask)
            .join(MoleculeState, ComputationTask.state_id == MoleculeState.id)
            .join(Molecule, MoleculeState.molecule_id == Molecule.id)
            .where(Molecule.project_name == project_name)
        ).one()
    return (molecules, tasks, str(latest))


def invalidate_cache(project_name: Optional[str] = None) -> None:
    """Drop cached analyses (all of them, or one project's)."""
    if project_name is None:
        _ANALYSIS_CACHE.clear()
    else:
        _ANALYSIS_CACHE.pop(project_name, None)


def analyze_project(project_name: str, use_cache: bool = True) -> dict:
    """Run state analysis for every molecule in a project.

    Results are cached per project. Every call re-reads and re-parses each
    molecule's ORCA outputs, which on a few thousand molecules is tens of
    thousands of file reads over a network mount inside a synchronous
    request handler -- enough to stall the controller for minutes when
    someone opens the State Analysis tab. The cache is keyed on a
    fingerprint of the project's task table, so new results invalidate it
    automatically.
    """
    if use_cache:
        signature = _cache_signature(project_name)
        cached = _ANALYSIS_CACHE.get(project_name)
        if cached is not None and cached[0] == signature:
            return cached[1]

    payload = _analyze_project_uncached(project_name)
    if use_cache:
        _ANALYSIS_CACHE[project_name] = (signature, payload)
    return payload


def _analyze_project_uncached(project_name: str) -> dict:
    """Run state analysis for every molecule in a project.

    For live projects, energies are read from the on-disk ORCA outputs
    via :class:`PipelineExtractor`. For archived projects (where
    comp_data/ has been wiped) we reconstruct the same data from the
    archive CSV written by :meth:`PipelineExtractor.archive_project`.
    The DB still holds every geometry row, so Kabsch RMSD works in both
    cases.
    """
    # Detect archived status first so we know which extraction path to use.
    with get_session() as _s:
        archived_count = _s.exec(
            select(func.count())
            .select_from(Molecule)
            .where(
                Molecule.project_name == project_name,
                Molecule.archived == True,  # noqa: E712
            )
        ).one()
        total_count = _s.exec(
            select(func.count())
            .select_from(Molecule)
            .where(Molecule.project_name == project_name)
        ).one()
    is_archived = total_count > 0 and archived_count == total_count

    all_results: list[ConformerResult] = []
    archive_source: Optional[str] = None
    if is_archived:
        # Resolve the archive CSV via the active Settings so it works
        # whether the dashboard or a CLI tool is the caller.
        try:
            from autodft.api.routes import get_active_settings
            settings = get_active_settings()
            csv_path = settings.export_data_path / project_name / f"{project_name}.csv"
        except Exception:
            csv_path = None
        if csv_path and csv_path.is_file():
            all_results = _load_results_from_archive_csv(csv_path)
            archive_source = str(csv_path)
    if not all_results:
        extractor = PipelineExtractor(project_name)
        all_results = extractor.extract_results(all_conformers=True)

    # (mol_id, state_desc) -> [ConformerResult, ...], ordered by opt_task_id
    groups: dict[tuple[int, str], list[ConformerResult]] = {}
    for r in all_results:
        groups.setdefault((r.molecule_id, r.state), []).append(r)
    for k in groups:
        groups[k].sort(key=lambda r: r.opt_task_id)

    with get_session() as session:
        mols = session.exec(
            select(Molecule)
            .where(Molecule.project_name == project_name)
            .order_by(col(Molecule.id).asc())
        ).all()
        mol_ids = [m.id for m in mols if m.id is not None]
        if not mol_ids:
            return {
                "project": project_name,
                "solvation_mecn": False,
                "sce_ref": SCE_ABS_MECN,
                "archived": is_archived,
                "archive_source": archive_source,
                "molecules": [],
            }

        # MeCN detection: scan every opt / SP header touched by this project.
        states = session.exec(
            select(MoleculeState).where(col(MoleculeState.molecule_id).in_(mol_ids))
        ).all()
        header_ids = {
            hid for st in states
            for hid in (st.optimization_header_id, st.singlepoint_header_id)
            if hid is not None
        }
        headers = (
            session.exec(
                select(ComputationHeader).where(col(ComputationHeader.id).in_(list(header_ids)))
            ).all()
            if header_ids else []
        )
        solvation_mecn = any(detects_mecn_solvation(h.header_text) for h in headers)

        # ...but decide per molecule when emitting potentials. A project-wide
        # any() meant one solvated header anywhere switched E vs SCE on for
        # every molecule in the project, including gas-phase ones, whose
        # "potential" would then be referenced against a solvated SCE scale.
        _header_by_id = {h.id: h for h in headers}
        mecn_by_molecule: dict[int, bool] = {}
        for st in states:
            st_mecn = any(
                detects_mecn_solvation(_header_by_id[hid].header_text)
                for hid in (st.optimization_header_id, st.singlepoint_header_id)
                if hid in _header_by_id
            )
            mecn_by_molecule[st.molecule_id] = (
                mecn_by_molecule.get(st.molecule_id, False) or st_mecn
            )

        opt_task_ids = [r.opt_task_id for r in all_results]
        opt_tasks = (
            session.exec(
                select(ComputationTask).where(col(ComputationTask.id).in_(opt_task_ids))
            ).all()
            if opt_task_ids else []
        )
        opt_by_id = {t.id: t for t in opt_tasks}
        out_geom_ids = [t.output_geometry_id for t in opt_tasks if t.output_geometry_id]
        out_geoms = (
            session.exec(
                select(MoleculeGeometry).where(col(MoleculeGeometry.id).in_(out_geom_ids))
            ).all()
            if out_geom_ids else []
        )
        geom_by_id = {g.id: g for g in out_geoms}

    def _coords_for(opt_task_id: int) -> Optional[np.ndarray]:
        task = opt_by_id.get(opt_task_id)
        if not task or not task.output_geometry_id:
            return None
        g = geom_by_id.get(task.output_geometry_id)
        if not g:
            return None
        _, coords = parse_xyz(g.xyz_data)
        return coords if coords.size else None

    def _rmsd_against(opt_task_id: int, ref: Optional[np.ndarray]) -> Optional[float]:
        if ref is None or ref.size == 0:
            return None
        coords = _coords_for(opt_task_id)
        if coords is None or coords.shape != ref.shape:
            return None
        try:
            return kabsch_rmsd(coords, ref)
        except Exception:
            return None

    def _pick_from(state: str, r: ConformerResult, ref: Optional[np.ndarray]) -> StatePick:
        return StatePick(
            state=state,
            conformer_index=r.conformer_index,
            opt_task_id=r.opt_task_id,
            rmsd_to_s0=(_rmsd_against(r.opt_task_id, ref) if state != "S0" else None),
            e_sp=r.e_singlepoint,
            e_correction=r.e_correction,
            e_combined=r.e_combined,
            e_vert_ox=r.e_vert_ox,
            e_vert_red=r.e_vert_red,
            e_vert_spin_change=r.e_vert_spin_change,
        )

    out_mols = []
    for m in mols:
        mol_id = m.id
        per_state: dict[str, list[ConformerResult]] = {}
        for (mid, st_desc), rs in groups.items():
            if mid != mol_id:
                continue
            per_state[st_desc] = rs

        if "S0" not in per_state:
            out_mols.append({
                "id": mol_id, "smiles": m.smiles, "modes": None,
                "note": "No S0 conformers extracted yet — analysis skipped.",
            })
            continue

        s0_lowest = min(
            (r for r in per_state["S0"] if r.e_combined is not None),
            key=lambda r: r.e_combined,
            default=None,
        )
        if s0_lowest is None:
            out_mols.append({
                "id": mol_id, "smiles": m.smiles, "modes": None,
                "note": "S0 has no conformer with both SP energy and free-energy correction.",
            })
            continue

        s0_ref_coords = _coords_for(s0_lowest.opt_task_id)

        # Mode 1 — lowest-energy conformer per state
        picks_low: dict[str, StatePick] = {}
        for st, pool in per_state.items():
            cand = min(
                (r for r in pool if r.e_combined is not None),
                key=lambda r: r.e_combined,
                default=None,
            )
            if cand is None:
                continue
            picks_low[st] = _pick_from(st, cand, s0_ref_coords)
        mol_mecn = mecn_by_molecule.get(mol_id, False)
        ana_low = _compute_derived(picks_low, mol_mecn)
        ana_low.mode = "lowest_energy"

        # Mode 2 — RMSD-matched conformer per state (S0 keeps its lowest)
        picks_rmsd: dict[str, StatePick] = {}
        if "S0" in picks_low:
            picks_rmsd["S0"] = picks_low["S0"]
        for st, pool in per_state.items():
            if st == "S0":
                continue
            scored: list[tuple[float, ConformerResult]] = []
            for r in pool:
                if r.e_combined is None:
                    continue
                rmsd = _rmsd_against(r.opt_task_id, s0_ref_coords)
                if rmsd is None:
                    continue
                scored.append((rmsd, r))
            if scored:
                scored.sort(key=lambda x: x[0])
                picks_rmsd[st] = _pick_from(st, scored[0][1], s0_ref_coords)
            elif st in picks_low:
                # RMSD couldn't be computed (e.g. missing geom) — fall back
                # to the lowest-energy pick so derived quantities still
                # populate. Tagged via rmsd_to_s0 = None.
                picks_rmsd[st] = picks_low[st]
        ana_rmsd = _compute_derived(picks_rmsd, mol_mecn)
        ana_rmsd.mode = "rmsd_matched"

        out_mols.append({
            "id": mol_id,
            "smiles": m.smiles,
            "modes": {
                "lowest_energy": _ana_to_dict(ana_low),
                "rmsd_matched":  _ana_to_dict(ana_rmsd),
            },
        })

    return {
        "project": project_name,
        "solvation_mecn": solvation_mecn,
        "sce_ref": SCE_ABS_MECN,
        "archived": is_archived,
        "archive_source": archive_source,
        "molecules": out_mols,
    }


def _ana_to_dict(a: ModeAnalysis) -> dict:
    return {
        "mode": a.mode,
        "picks": {
            st: {
                "state": p.state,
                "conformer_index": p.conformer_index,
                "opt_task_id": p.opt_task_id,
                "rmsd_to_s0": p.rmsd_to_s0,
                "e_sp": p.e_sp,
                "e_correction": p.e_correction,
                "e_combined": p.e_combined,
                "e_vert_ox": p.e_vert_ox,
                "e_vert_red": p.e_vert_red,
                "e_vert_spin_change": p.e_vert_spin_change,
            }
            for st, p in a.picks.items()
        },
        "triplet": {
            "dE_adiabatic_h": a.triplet_dE_adiabatic_h,
            "lambda_fwd_h":   a.triplet_lambda_fwd_h,
            "lambda_bwd_h":   a.triplet_lambda_bwd_h,
        },
        "ox": {
            "dG_h":          a.ox_dG_h,
            "E_vs_sce_v":    a.ox_E_vs_sce_v,
            "lambda_fwd_h":  a.ox_lambda_fwd_h,
            "lambda_bwd_h":  a.ox_lambda_bwd_h,
        },
        "red": {
            "dG_h":          a.red_dG_h,
            "E_vs_sce_v":    a.red_E_vs_sce_v,
            "lambda_fwd_h":  a.red_lambda_fwd_h,
            "lambda_bwd_h":  a.red_lambda_bwd_h,
        },
    }


# ----------------------------------------------------------------------
# XLSX export
# ----------------------------------------------------------------------


def _xlsx_mode_columns() -> list[tuple[str, str]]:
    """Per-mode sheet column definitions: (header, key path)."""
    return [
        ("mol_id",               "id"),
        ("smiles",               "smiles"),
        # Conformer picks per state (index + RMSD vs S0 in Å)
        ("S0_conf",              "picks.S0.conformer_index"),
        ("T1_conf",              "picks.T1.conformer_index"),
        ("T1_rmsd_to_S0_A",      "picks.T1.rmsd_to_s0"),
        ("ox_conf",              "picks.ox.conformer_index"),
        ("ox_rmsd_to_S0_A",      "picks.ox.rmsd_to_s0"),
        ("red_conf",             "picks.red.conformer_index"),
        ("red_rmsd_to_S0_A",     "picks.red.rmsd_to_s0"),
        # Triplet (Hartree)
        ("triplet_dE_adiab_Eh",  "triplet.dE_adiabatic_h"),
        ("triplet_lambda_fwd_Eh", "triplet.lambda_fwd_h"),
        ("triplet_lambda_bwd_Eh", "triplet.lambda_bwd_h"),
        # Oxidation
        ("ox_dG_Eh",             "ox.dG_h"),
        ("ox_E_vs_SCE_V",        "ox.E_vs_sce_v"),
        ("ox_lambda_fwd_Eh",     "ox.lambda_fwd_h"),
        ("ox_lambda_bwd_Eh",     "ox.lambda_bwd_h"),
        # Reduction
        ("red_dG_Eh",            "red.dG_h"),
        ("red_E_vs_SCE_V",       "red.E_vs_sce_v"),
        ("red_lambda_fwd_Eh",    "red.lambda_fwd_h"),
        ("red_lambda_bwd_Eh",    "red.lambda_bwd_h"),
    ]


def _lookup(obj: object, path: str):
    cur = obj
    for part in path.split("."):
        if cur is None:
            return None
        if isinstance(cur, dict):
            cur = cur.get(part)
        else:
            cur = getattr(cur, part, None)
    return cur


def build_xlsx_bytes(payload: dict) -> bytes:
    """Render the state-analysis ``payload`` (the dict returned by
    :func:`analyze_project`) into an XLSX file and return the raw bytes.

    Layout:
        - "Summary"        — project name + MeCN flag + SCE ref + archived flag
        - "Lowest Energy"  — one row per molecule, every metric (Hartree / V)
        - "RMSD Matched"   — same columns, different conformer picks
        - "Conformers"     — every conformer's energies for both modes
                             (S0/T1/ox/red with the SP / correction /
                             combined / vert_* fields in Hartree)
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F3340")

    # ── Summary sheet ────────────────────────────────────────────────
    summary["A1"] = "Project"
    summary["B1"] = payload.get("project")
    summary["A2"] = "Archived"
    summary["B2"] = "yes" if payload.get("archived") else "no"
    summary["A3"] = "MeCN solvation"
    summary["B3"] = "detected" if payload.get("solvation_mecn") else "not detected"
    summary["A4"] = "SCE reference (V)"
    summary["B4"] = payload.get("sce_ref")
    summary["A5"] = "SCE convention"
    summary["B5"] = "Pavlishchuk & Addison, Inorg. Chim. Acta 2000"
    summary["A6"] = "Source"
    summary["B6"] = payload.get("archive_source") or "live ORCA outputs"
    summary["A7"] = "Molecules"
    summary["B7"] = len(payload.get("molecules", []))
    for row in range(1, 8):
        summary.cell(row=row, column=1).font = Font(bold=True)
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 60

    # ── Per-mode summary sheets ──────────────────────────────────────
    columns = _xlsx_mode_columns()
    for mode_key, sheet_title in (
        ("lowest_energy", "Lowest Energy"),
        ("rmsd_matched",  "RMSD Matched"),
    ):
        ws = wb.create_sheet(sheet_title)
        for col_idx, (label, _path) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r_idx, m in enumerate(payload.get("molecules", []), start=2):
            ana = (m.get("modes") or {}).get(mode_key)
            row_src = {**(ana or {}), "id": m.get("id"), "smiles": m.get("smiles")}
            for c_idx, (_label, path) in enumerate(columns, start=1):
                if path in ("id", "smiles"):
                    val = m.get(path)
                else:
                    val = _lookup(row_src, path) if ana else None
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Sensible default widths.
        for col_letter, width in (
            ("A", 8), ("B", 30), ("C", 8), ("D", 8), ("E", 14),
            ("F", 8), ("G", 14), ("H", 8), ("I", 14),
        ):
            ws.column_dimensions[col_letter].width = width
        for col_idx in range(10, len(columns) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

    # ── Conformers sheet (raw per-conformer energies, both modes) ────
    ws = wb.create_sheet("Conformers")
    headers = [
        "mode", "mol_id", "smiles", "state", "conformer_index", "opt_task_id",
        "rmsd_to_S0_A", "e_sp_Eh", "e_correction_Eh", "e_combined_Eh",
        "e_vert_ox_Eh", "e_vert_red_Eh", "e_vert_spin_change_Eh",
    ]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    r = 2
    for m in payload.get("molecules", []):
        modes = m.get("modes") or {}
        for mode_key in ("lowest_energy", "rmsd_matched"):
            ana = modes.get(mode_key)
            if not ana:
                continue
            picks = ana.get("picks") or {}
            for st, p in picks.items():
                ws.cell(row=r, column=1, value=mode_key)
                ws.cell(row=r, column=2, value=m.get("id"))
                ws.cell(row=r, column=3, value=m.get("smiles"))
                ws.cell(row=r, column=4, value=st)
                ws.cell(row=r, column=5, value=p.get("conformer_index"))
                ws.cell(row=r, column=6, value=p.get("opt_task_id"))
                ws.cell(row=r, column=7, value=p.get("rmsd_to_s0"))
                ws.cell(row=r, column=8, value=p.get("e_sp"))
                ws.cell(row=r, column=9, value=p.get("e_correction"))
                ws.cell(row=r, column=10, value=p.get("e_combined"))
                ws.cell(row=r, column=11, value=p.get("e_vert_ox"))
                ws.cell(row=r, column=12, value=p.get("e_vert_red"))
                ws.cell(row=r, column=13, value=p.get("e_vert_spin_change"))
                r += 1
    for col_idx, width in enumerate(
        [16, 8, 30, 8, 8, 10, 14, 18, 18, 18, 18, 18, 22], start=1
    ):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
